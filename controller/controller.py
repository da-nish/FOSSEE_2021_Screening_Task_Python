from PyQt5.QtCore import pyqtSignal, QObject
from openpyxl import load_workbook

class Controller(QObject):
    """docstring for Controller"""
    update_signal = pyqtSignal(int)

    def __init__(self, model):
        super(Controller, self).__init__()


        self._model = model
        self.SHOW_CONSOLE = False  # Toggle console output
        self.header = []
        self.record = []
        self.excel_IDs = []
        self.current_section_append = 1
        self.current_section_display = 1
        # mapping ui selected index with excel sheet index { UI_index: excel_sheet_index }
        self.sheet_combobox_index = {0: 1, 1: 0, 2: 2}
        self.wb = load_workbook(filename="new_sections.xlsx")
        self.change_append(0)
        self.change_display(0)

    def change_display(self, index):

        self.record, self.header = self._model.getRecord(index)
        self.current_section_display = index


    def change_append(self, index):
        if self.current_section_append == index:
            return
        try:
            self.wb.active = self.sheet_combobox_index.get(index)
            sheet = self.wb.active
            self.current_section_append = index

            ignore_first = True
            tempID = []
            for row in sheet.rows:
                if ignore_first:  # ignoring excel header row
                    ignore_first = False
                    continue
                tempID.append(str(int(row[0].value)))
        except Exception as e:
            print("//",e)
        self.excel_IDs = tempID

    # check - ui selected id with section
    def is_exist_id(self, index, id):

        self.wb.active = self.sheet_combobox_index.get(index)
        sheet = self.wb.active
        self.console([sheet])

        ignore_first = True
        for row in sheet.iter_rows(max_col = 24):
            if ignore_first:  # ignoring excel header row
                ignore_first = False
                continue

            val = str(int(row[0].value))
            try:
                if val == id:
                    return True, row  # returning first occurrence
            except Exception as e:
                return False, str(e)

        return False, "Not Exist"

    def append_data(self, index, id):
        isexit, row = self.is_exist_id(index, id)
        if not isexit:
            self.console([row])
            self.console(['Error:', row, '(on is_exist_id method)'])

        # converting row object to string list
        record = []
        for r in row:
            try:
                record.append(r.value)
            except Exception as e:
                print(r, e)
                if r.value == '':
                    record.append(None)
                else:
                    record.append(r)

        record.pop(0) # removing id
        try:
            if self._model.append_record(index, record):
                self.console(["Append: Success !!"])
                self.change_display(index)
                self.update_signal.emit(index)

            else:
                self.console(["Append: Fail !!"])
        except Exception as e:
            self.console(['Error:', e, '(on append)'])
            # return True


    def console(self, arg):
        if self.SHOW_CONSOLE is False:
            return

        for msg in arg:
            print(msg, ' ', end='')
        print()
